import openpyxl

ROW_NUM = 169
COL_NUM = 19


file_name ='tuan'

wb = openpyxl.load_workbook('source/'+file_name+'.xlsx')
sheet_list = wb.get_sheet_names()
space = 0

des_wb = openpyxl.Workbook()
des_sheet = des_wb.active


sheet = wb.get_sheet_by_name(sheet_list[0])

space=0

dr = 2
for dong in range(2,10):
    for r in range(dong, ROW_NUM+1, 8):
        country = sheet.cell(row=r, column=1).value
        var = sheet.cell(row=r, column=2).value
        for c in range(3, COL_NUM+1):
            value = sheet.cell(row=r, column=c).value
            year = sheet.cell(row=1, column=c).value
            des_sheet.cell(row=1, column=3+space).value = var
            des_sheet.cell(row=dr, column=1+space).value = country
            des_sheet.cell(row=dr, column=2+space).value = year
            des_sheet.cell(row=dr, column=3+space).value = value

            dr+=1      
    space+=3
    dr=2
des_wb.save('dx/tuan.xlsx')