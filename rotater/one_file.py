import openpyxl

ROW_NUM = 652
COL_NUM = 6


file_name ='bien2.xlsx'

wb = openpyxl.load_workbook('src/'+file_name)
sheet_list = wb.get_sheet_names()
for sheet_name in sheet_list:
    sheet = wb.get_sheet_by_name(sheet_name)

    des_wb = openpyxl.Workbook()
    des_sheet = des_wb.active

    dr = 1

    for r in range(2, ROW_NUM+1):
        symbol = sheet.cell(row=r, column=1).value
        for c in range(2, COL_NUM+1):
            value = sheet.cell(row=r, column=c).value
            year = sheet.cell(row=1, column=c).value
            
            des_sheet.cell(row=dr, column=1).value = symbol
            des_sheet.cell(row=dr, column=2).value = year
            des_sheet.cell(row=dr, column=3).value = value

            dr+=1
       
    des_wb.save('dx/'+sheet_name+'.xlsx')

    print(sheet_name)