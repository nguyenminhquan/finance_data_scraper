import openpyxl

ROW_NUM = 1706
YEAR_NUM = 5    


file_name ='cashflow'

wb = openpyxl.load_workbook('source/'+file_name+'.xlsx')
sheet_list = wb.get_sheet_names()
space = 0

des_wb = openpyxl.Workbook()
des_sheet = des_wb.active

for sheet_name in sheet_list:
    sheet = wb.get_sheet_by_name(sheet_name)


    dc = 2
    for y in range(2, YEAR_NUM+2):
        des_sheet.cell(row=1, column=dc).value = sheet.cell(row=y, column=2).value
        dc+=1

    dr = 2
    for r in range(2, ROW_NUM+1, YEAR_NUM):
        des_sheet.cell(row=dr, column=1).value = sheet.cell(row=r, column=1).value
        dc = 2
        for v in range(r, r+YEAR_NUM):

            des_sheet.cell(row=dr, column=dc).value = sheet.cell(row=v, column=3).value
            dc+=1

        dr+=1
    print(sheet_name)   
des_wb.save('dx/'+file_name+'.xlsx')