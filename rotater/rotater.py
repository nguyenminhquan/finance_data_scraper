import openpyxl

ROW_NUM = 254
COL_NUM = 6

LIST_FILE = 'list.txt'

file_list = []
with open(LIST_FILE, 'r') as f:
    for line in f:
        file_list.append(line.strip())

for file_name in file_list:
    wb = openpyxl.load_workbook('src/'+file_name+'.xlsx')
    sheet_list = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheet_list[0])

    des_wb = openpyxl.Workbook()
    des_sheet = des_wb.active

    dr = 1

    for r in range(2, ROW_NUM+1):
        symbol = sheet.cell(row=r, column=1).value
        for c in range(2, COL_NUM+1):
            value = sheet.cell(row=r, column=c).value
            year = sheet.cell(row=1, column=c).value
            
            des_sheet.cell(row=dr, column=1).value = symbol
            des_sheet.cell(row=dr, column=2).value = year
            des_sheet.cell(row=dr, column=3).value = value

            dr+=1
       
    des_wb.save('result/'+file_name+'_dx.xlsx')
    wb.save(file_name+'.xlsx')
    print(file_name)
