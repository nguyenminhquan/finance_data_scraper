import openpyxl

ROW_NUM = 339
COL_NUM = 6

file_name ='tonghop'

wb = openpyxl.load_workbook('src/'+file_name+'.xlsx')
sheet_list = wb.get_sheet_names()
space = 0

des_wb = openpyxl.Workbook()
des_sheet = des_wb.active

for sheet_name in sheet_list:
    sheet = wb.get_sheet_by_name(sheet_name)

    dr = 2
    
    for r in range(2, ROW_NUM+1):
        symbol = sheet.cell(row=r, column=1).value
        for c in range(2, COL_NUM+1):
            value = sheet.cell(row=r, column=c).value
            year = sheet.cell(row=1, column=c).value
            
            des_sheet.cell(row=dr, column=1+space).value = symbol
            des_sheet.cell(row=dr, column=2+space).value = year
            des_sheet.cell(row=dr, column=3+space).value = value

            dr+=1
    des_sheet.cell(row=1, column=3+space).value = sheet_name      
    space+=3
    print(sheet_name)   
des_wb.save('result/'+file_name+'.xlsx')